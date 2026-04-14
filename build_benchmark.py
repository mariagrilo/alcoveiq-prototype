from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Sheet 1: Feature Comparison ──
ws = wb.active
ws.title = "Feature Comparison"

thin = Side(style='thin', color='D0D0D0')
border = Border(top=thin, left=thin, right=thin, bottom=thin)

hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill('solid', fgColor='1e3a5f')
cat_font = Font(name='Arial', bold=True, size=11, color='1e3a5f')
cat_fill = PatternFill('solid', fgColor='e8edf3')
body_font = Font(name='Arial', size=10)
yes_fill = PatternFill('solid', fgColor='e8f5e9')
no_fill = PatternFill('solid', fgColor='fce4ec')
partial_fill = PatternFill('solid', fgColor='fff8e1')
harbor_adv_fill = PatternFill('solid', fgColor='e3f2fd')

headers = ['Category', 'Feature', 'Harbor Compliance', 'AlcoveIQ', 'AlcoveIQ Advantage']
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 42
ws.column_dimensions['E'].width = 48

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

ws.freeze_panes = 'A2'
ws.auto_filter.ref = 'A1:E1'

data = [
    # (category, feature, harbor, alcoveiq, advantage)
    ('Dashboard', None, None, None, None),
    ('', 'At-a-glance compliance health score', 'No', 'Yes — 94% with color coding', 'Immediate situational awareness'),
    ('', 'Portfolio summary stats', 'No', 'Yes — entity count, deadlines, fees', 'Quantified portfolio health'),
    ('', 'Attention/alert banner', 'Partial — generic "5 gaps detected"', 'Yes — specific entity names + issues', 'Actionable, not generic'),
    ('', 'Upcoming filings widget', 'No — just a count', 'Yes — calendar-style with dates + badges', 'Visual deadline urgency'),
    ('', 'AI assistant on dashboard', 'No — chat is sales-oriented', 'Yes — contextual prompts, custom input', 'Functional, not promotional'),
    ('', 'Quick order creation', 'No', 'Yes — "+ New Order" always visible', 'One-click access'),

    ('Entity Management', None, None, None, None),
    ('', 'Entity list with status indicators', 'No — basic table, 4 columns', 'Yes — type badges, RA status, compliance, deadlines', 'Rich status visibility at a glance'),
    ('', 'Entity detail view', 'Full-page with 6 tabs (Details, Locations, Financials, Key Personnel, Signers, Records)', 'Slide-in sidebar with summary, officers, compliance, orders', 'Context-preserving, no page navigation'),
    ('', 'Entity type categorization', 'No', 'Yes — Funds, SPVs, GP Entities, HoldCos filter chips', 'Portfolio-aware grouping for PE/VC'),
    ('', 'Multi-state registration map', 'Yes — US heat map with registration density', 'No', '— Harbor advantage'),
    ('', 'Company Groups', 'Yes', 'No', '— Harbor advantage'),
    ('', 'Foreign qualification tracking', 'Yes — Registration List', 'Yes — foreign states column', 'Feature parity'),
    ('', 'Entity search & filtering', 'Basic — name/DBA search only', 'Yes — search + type chips + sortable columns', 'More filter dimensions'),
    ('', 'Org chart visualization', 'No', 'Yes — hierarchical ownership view', 'Visual ownership structure'),

    ('Compliance & Filings', None, None, None, None),
    ('', 'Compliance calendar', 'Yes — flat table, no urgency indicators', 'Yes — color-coded badges, stat cards, filters', 'Visual urgency system'),
    ('', 'Overdue / due soon indicators', 'No — all items look identical', 'Yes — red/amber/green/blue status system', 'Clear priority signaling'),
    ('', 'Filing requirements detail', 'No', 'Yes — frequency, agency, penalty info', 'Contextual filing information'),
    ('', 'Blocker warnings (e.g. lapsed RA)', 'No', 'Yes — lapsed RA blocks filing alerts', 'Prevents failed filings'),
    ('', '"File Now" action button', 'No — "Order managed service" sales redirect', 'Yes — direct action on every non-auto-filed item', 'Self-service filing'),
    ('', 'Auto-filed status tracking', 'Not visible', 'Yes — Auto-filed badge', 'Automation visibility'),
    ('', 'Compliance stat cards', 'No', 'Yes — overdue, due soon, on track, auto-filed counts', 'Quantified compliance posture'),
    ('', 'Multi-dimensional compliance filters', 'Partial — Assigned To, Date Due only', 'Yes — search, entity, state, status chips', 'AND-combining filter pattern'),

    ('Orders & Projects', None, None, None, None),
    ('', 'Order/project list', 'Yes — sparse (Project, Status columns only)', 'Yes — 27+ orders with service, entity, state, priority, progress', 'Information-rich order tracking'),
    ('', 'Order progress visualization', 'No', 'Yes — progress bars with percentage', 'Visual progress tracking'),
    ('', 'Order detail view', 'Not observed', 'Yes — slide-in sidebar with timeline and steps', 'Detailed order lifecycle view'),
    ('', 'Order creation flow', 'Sales-mediated (managed services)', 'Self-service modal with service type grid', 'Direct self-service ordering'),
    ('', 'Service type filtering', 'Status filters only', 'Yes — Formation, Annual Report, Foreign Qual, etc.', 'Service-specific views'),
    ('', 'Priority levels (Rush, Urgent)', 'Not visible', 'Yes — Standard, Rush, Urgent, Same Day', 'Expedited filing options'),

    ('AI & Automation', None, None, None, None),
    ('', 'AI chat assistant', 'Harbor Copilot — generic, sales-oriented', 'AlcoveIQ Assistant — entity-aware, can file orders', 'Functional AI vs lead-gen chatbot'),
    ('', 'Entity lookup via AI', 'No', 'Yes — full entity cards with officers, compliance, RA', 'Deep data access via natural language'),
    ('', 'Order creation via AI', 'No', 'Yes — conversational order flow with confirmation', 'Natural language ordering'),
    ('', 'Fee breakdown via AI', 'No', 'Yes — itemized filing fees with totals', 'Financial visibility on demand'),
    ('', 'Proactive compliance alerts', 'Generic — "5 gaps detected"', 'Specific — entity names, blocker chains, sequences', 'Actionable intelligence'),
    ('', 'Recent queries memory', 'No', 'Yes — recent query chips for quick re-access', 'Usage-aware UX'),

    ('Documents', None, None, None, None),
    ('', 'Document vault / storage', 'Yes — Records Manager', 'Yes — Documents page with filters', 'Feature parity'),
    ('', 'Document categorization', 'Yes — by type', 'Yes — Formation, Compliance, Legal, Tax, Correspondence', 'Feature parity'),
    ('', 'Document search & filtering', 'Not observed', 'Yes — search + entity + type filters', 'Multi-filter document search'),

    ('UX & Design', None, None, None, None),
    ('', 'Visual design era', '2018-era enterprise SaaS (light, blue/white/orange)', 'Modern dark mode (navy/teal, semantic colors)', 'Contemporary aesthetic'),
    ('', 'Navigation complexity', '15+ sidebar items across 6 sections', '7 items + actions section', 'Reduced cognitive load'),
    ('', 'Information density', 'Low — single-purpose pages, lots of whitespace', 'High — multi-data dashboards, sidebar panels', 'More information per screen'),
    ('', 'Detail view pattern', 'Full page navigation (loses context)', 'Slide-in sidebar panels (stays in context)', 'Context preservation'),
    ('', 'Keyboard shortcuts', 'None observed', 'Yes — ⌘K for assistant', 'Power user support'),

    ('Data Depth & Breadth', None, None, None, None),
    ('', 'Requirements Research Engine', 'Yes', 'No', '— Harbor advantage'),
    ('', 'License Manager', 'Yes', 'No', '— Harbor advantage'),
    ('', 'Tax Manager', 'Yes', 'No', '— Harbor advantage'),
    ('', 'DBA tracking', 'Yes — DBA List tab', 'No', '— Harbor advantage'),
    ('', 'Billing / invoicing', 'Yes — integrated billing', 'No — prototype stage', '— Harbor advantage (mature billing)'),
    ('', 'Multi-user / team support', 'Yes — Users settings', 'Partial — single user prototype', '— Harbor advantage (team features)'),
]

row = 2
for item in data:
    cat, feat, harbor, alcove, adv = item
    if feat is None:  # Category row
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = cat_fill
            cell.font = cat_font
            cell.border = border
        ws.cell(row=row, column=1, value=cat)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1).alignment = Alignment(vertical='center')
    else:
        ws.cell(row=row, column=1, value=cat).font = body_font
        ws.cell(row=row, column=2, value=feat).font = body_font
        
        hc = ws.cell(row=row, column=3, value=harbor)
        hc.font = body_font
        if harbor and harbor.startswith('No'):
            hc.fill = no_fill
        elif harbor and harbor.startswith('Yes'):
            hc.fill = yes_fill
        elif harbor and harbor.startswith('Partial'):
            hc.fill = partial_fill
        
        ac = ws.cell(row=row, column=4, value=alcove)
        ac.font = body_font
        if alcove and alcove.startswith('No'):
            ac.fill = no_fill
        elif alcove and alcove.startswith('Yes'):
            ac.fill = yes_fill
        elif alcove and alcove.startswith('Partial'):
            ac.fill = partial_fill
        
        adv_cell = ws.cell(row=row, column=5, value=adv)
        adv_cell.font = body_font
        if adv and adv.startswith('—'):
            adv_cell.fill = harbor_adv_fill
        
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
            ws.cell(row=row, column=c).alignment = Alignment(vertical='center', wrap_text=True)
    row += 1

# ── Sheet 2: Summary Scorecard ──
ws2 = wb.create_sheet("Summary Scorecard")

for c, h in enumerate(['Category', 'Harbor Compliance', 'AlcoveIQ', 'Winner'], 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 18

scores = [
    ('Dashboard', '★★☆☆☆', '★★★★★', 'AlcoveIQ'),
    ('Entity Management', '★★★☆☆', '★★★★☆', 'AlcoveIQ'),
    ('Compliance Tracking', '★★☆☆☆', '★★★★★', 'AlcoveIQ'),
    ('Orders / Projects', '★★☆☆☆', '★★★★☆', 'AlcoveIQ'),
    ('AI Integration', '★☆☆☆☆', '★★★★★', 'AlcoveIQ'),
    ('Documents', '★★★☆☆', '★★★☆☆', 'Tie'),
    ('Visual Design', '★★☆☆☆', '★★★★☆', 'AlcoveIQ'),
    ('Data Depth & Breadth', '★★★★☆', '★★★☆☆', 'Harbor'),
    ('Multi-state Visualization', '★★★★☆', '★★☆☆☆', 'Harbor'),
]

alcove_win = PatternFill('solid', fgColor='e8f5e9')
harbor_win = PatternFill('solid', fgColor='e3f2fd')
tie_fill = PatternFill('solid', fgColor='f5f5f5')

for i, (cat, hc, aq, winner) in enumerate(scores, 2):
    ws2.cell(row=i, column=1, value=cat).font = Font(name='Arial', bold=True, size=10)
    ws2.cell(row=i, column=2, value=hc).font = Font(name='Arial', size=12)
    ws2.cell(row=i, column=2).alignment = Alignment(horizontal='center')
    ws2.cell(row=i, column=3, value=aq).font = Font(name='Arial', size=12)
    ws2.cell(row=i, column=3).alignment = Alignment(horizontal='center')
    w = ws2.cell(row=i, column=4, value=winner)
    w.font = Font(name='Arial', bold=True, size=10)
    w.alignment = Alignment(horizontal='center')
    
    fill = alcove_win if winner == 'AlcoveIQ' else harbor_win if winner == 'Harbor' else tie_fill
    for c in range(1, 5):
        ws2.cell(row=i, column=c).fill = fill
        ws2.cell(row=i, column=c).border = border

r = len(scores) + 3
ws2.cell(row=r, column=1, value='Key Takeaways').font = Font(name='Arial', bold=True, size=13, color='1e3a5f')
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

takeaways = [
    '1. Harbor is a data repository; AlcoveIQ is a workflow tool that drives action.',
    '2. Harbor\'s AI ("Copilot") is sales-oriented; AlcoveIQ\'s AI is operational and entity-aware.',
    '3. AlcoveIQ excels at urgency communication, self-service actions, and context preservation.',
    '4. Harbor has deeper data modules (Tax Manager, License Manager, Requirements Research Engine).',
    '5. AlcoveIQ\'s sidebar panel pattern eliminates context-switching full-page loads.',
]

for j, t in enumerate(takeaways):
    cell = ws2.cell(row=r+1+j, column=1, value=t)
    cell.font = Font(name='Arial', size=10)
    ws2.merge_cells(start_row=r+1+j, start_column=1, end_row=r+1+j, end_column=4)

out = '/Users/mariagrilo/Documents/AlcoveIQ/AlcoveIQ_vs_HarborCompliance_Benchmark.xlsx'
wb.save(out)
print(f'Saved to {out}')
